import re
import openpyxl
'''
Regex operation for matching the correct email from employee list
'''
wb=openpyxl.load_workbook('Book1.xlsx')
sheet=wb.active
row=sheet.max_row
column=sheet.max_column
for i in range(1,row):
    cell_val=sheet.cell(row=i,column=3)
    email=cell_val.value
    if re.match(r'[a-z]*[0-9]*[A-Z]*@gmail\.com',str(email)) is not None:
        print(email)


