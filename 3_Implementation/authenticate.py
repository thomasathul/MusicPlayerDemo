"""
Authenticating the username and password module
"""
import re
from openpyxl import load_workbook


class AuthenticateAccount:
    """
     Class with authentication methods
    """

    def __init__(self, username=None, password=None):
        self.username = username
        self.password = password

    @staticmethod
    def __hashcode__(password):
        """
        Creating hashcode of the string
        :param password:
        :return: asciisum
        """
        letterlist = list(password)
        asciisum = 0
        for alpha in letterlist:
            asciisum += (ord(alpha) * letterlist.index(alpha))

        asciisum //= 5
        return asciisum

    def passwordhashing(self):
        """
        Calling abstract hashing method
        :return: hashedpassword
        """
        hashedpassword = self.__hashcode__(self.password)
        return hashedpassword


class SignUp(AuthenticateAccount):
    """
    Sign up page for the music streaming app
    """

    def usernamevalidity(self):
        """
        Checking username is valid or not
        :return: Status.SUCCESS.value or Status.ERROR.value
        """
        valid = bool(re.match(r'\b[A-Za-z0-9._%+-]+@[a-z0-9.-]+\.[a-z]{2,}\b', self.username))
        return valid

    def usernameexist(self):
        """
        Checking if username exists
        :return: Status.SUCCESS.value or Status.ERROR.value
        """
        database = load_workbook("database/database.xlsx")
        account_sheet = database["Sheet1"]
        s_row = account_sheet.max_row
        for number in range(1, s_row + 1):
            username_val = account_sheet.cell(row=number, column=1)
            if username_val.value == self.username:
                return False
        return True

    def passwordvalidity(self):
        """
        Checking the validity of the password
        :return:Status.SUCCESS.value or Status.ERROR.value
        """
        valid = bool(re.match(r'[\w]{8,}', self.password))
        return valid

    def saveindatabase(self):
        """
        Saving the username and password in database
        :return: Status.SUCCESS.value
        """
        database = load_workbook("database/database.xlsx")
        account_sheet = database["Sheet1"]
        s_row = account_sheet.max_row
        cell_val_username = account_sheet.cell(row=s_row + 1, column=1)
        cell_val_password = account_sheet.cell(row=s_row + 1, column=2)
        cell_val_username.value = self.username
        cell_val_password.value = self.passwordhashing()
        database.save("database/database.xlsx")
        return True  # "Account successfully created"


class LogIn(AuthenticateAccount):
    """
     Log in page for the music streaming app
    """

    def searchindatabase(self):
        """
        Searching the username and password in the database
        :return: Status.SUCCESS.value
        """
        if self.username is None and self.password is None:
            return False
        database = load_workbook("database/database.xlsx")
        account_sheet = database["Sheet1"]
        s_row = account_sheet.max_row
        for number in range(1, s_row+1):
            username_val = account_sheet.cell(row=number, column=1)
            if username_val.value == self.username:
                password_val = account_sheet.cell(row=number, column=2)
                if password_val.value == self.passwordhashing():
                    return True  # "Login Successful"
                return "Incorrect"  # "Wrong password"
        return False  # "Account doesn't exist
