"""
Defining all the Music Player features
"""
import re
import pygame
from openpyxl import load_workbook


def play(musicfile):
    """
    PLaying the audio file
    :param musicfile:
    :return: True
    """
    pygame.mixer.init()
    pygame.mixer.music.load("music/"+musicfile)
    pygame.mixer.music.play()
    return True


class MusicPlayer:
    """
    Music Player for selecting the songs
    """
    @staticmethod
    def selected_song(row_num):
        """
        Getting the selected song file name number
        :param row_num:
        :return: value
        """
        database = load_workbook("database/database.xlsx")
        music_sheet = database['Sheet2']
        file_name = music_sheet.cell(row=row_num, column=3)
        return file_name.value

    def select_song(self, track_name):
        """
        Getting the song file name
        :param track_name:
        :return: file_name or False
        """
        database = load_workbook("database/database.xlsx")
        music_sheet = database['Sheet2']
        s_row = music_sheet.max_row
        for number in range(1, s_row+1):
            music_val = music_sheet.cell(row=number, column=1)
            if re.match(track_name, music_val.value, re.I):
                return self.selected_song(number)
        return track_name


class SearchTracks(MusicPlayer):
    """
    Searching music tracks from the list
    """

    def __init__(self, trackname=None):
        self.trackname = trackname

    def searchkeyword(self):
        """
        Searching for the keyword
        :return: music and singer list
        """
        music_list = []
        singer_list = []
        database = load_workbook("database/database.xlsx")
        music_sheet = database['Sheet2']
        s_row = music_sheet.max_row
        for number in range(1, s_row+1):
            music_val = music_sheet.cell(row=number, column=1)
            singer_val = music_sheet.cell(row=number, column=2)
            searchval = re.search(self.trackname, music_val.value, re.I)
            if searchval is not None:
                music_list.append(music_val.value)
                singer_list.append(singer_val.value)
        if not music_list:
            return None
        return music_list, singer_list


class BrowseTracks(MusicPlayer):
    """
    Browsing music tracks from the list
    """
    @staticmethod
    def browse_list():
        """
        Browsing through the song list
        :return: music and singer list
        """
        music_list = []
        singer_list = []
        database = load_workbook("database/database.xlsx")
        music_sheet = database['Sheet2']
        s_row = music_sheet.max_row
        for number in range(1, s_row + 1):
            music_val = music_sheet.cell(row=number, column=1)
            singer_val = music_sheet.cell(row=number, column=2)
            if music_val.value != "":
                music_list.append(music_val.value)
                singer_list.append(singer_val.value)
        return music_list, singer_list
